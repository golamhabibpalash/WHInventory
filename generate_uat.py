#!/usr/bin/env python3
"""
uStock UAT Spreadsheet Generator
Run:   python3 generate_uat.py
Needs: pip install openpyxl
Out:   uStock_UAT_Testing.xlsx  (import into Google Sheets via File → Import)
"""

import sys, subprocess

# ── Auto-install openpyxl if missing ─────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl …")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────────────────────────────────────
NAVY       = "1F3864"
WHITE      = "FFFFFF"
LGRAY      = "F2F2F2"
MGRAY      = "D9D9D9"

PHASE_CLR = {
    "P1": "4472C4", "P2": "ED7D31", "P3": "70AD47",
    "P4": "00B0F0", "P5": "7030A0", "P6": "C00000",
    "P7": "FFC000", "P8": "595959",
    "BUG": "C00000", "REQ": "375623", "SIGN": "1F3864",
}

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _fill(c):    return PatternFill("solid", fgColor=c)
def _font(bold=False, sz=10, clr="000000"):
    return Font(bold=bold, size=sz, color=clr)
def _align(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, val, width, bg):
    c = ws.cell(row=row, column=col, value=val)
    c.fill  = _fill(bg)
    c.font  = _font(True, 10, WHITE)
    c.alignment = _align("center")
    c.border = _border()
    ws.column_dimensions[get_column_letter(col)].width = width

def _data_cell(ws, row, col, val, bg=WHITE, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill  = _fill(bg)
    c.font  = _font(bold, 10)
    c.alignment = _align(align)
    c.border = _border()
    return c

def _title(ws, text, bg, row=1, cols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg); c.font = _font(True, 14, WHITE)
    c.alignment = _align("center"); ws.row_dimensions[row].height = 28

def _section(ws, text, row, cols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(MGRAY); c.font = _font(True, 11, "000000")
    c.alignment = _align("left"); ws.row_dimensions[row].height = 22

def _dropdown(ws, col_letter, r1, r2, opts):
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(opts) + '"',
                        allow_blank=True)
    dv.sqref = f"{col_letter}{r1}:{col_letter}{r2}"
    ws.add_data_validation(dv)

def _header_row(ws, row, headers, widths, bg):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr_cell(ws, row, i, h, w, bg)
    ws.row_dimensions[row].height = 20

# ─────────────────────────────────────────────────────────────────────────────
# TEST CASE DATA
# ─────────────────────────────────────────────────────────────────────────────
# Each tuple: (Category, Test Case, Steps, Expected Result)

P1 = [
    ("Authentication",   "Admin login",                   "Go to /Accounts/Login → enter admin@root.com / 123456",                              "Dashboard loads successfully"),
    ("User Mgmt",        "Create Tester 2 account",       "Admin → Users → New → enter Sales Tester email & password",                         "User created, visible in user list"),
    ("User Mgmt",        "Create Tester 3 account",       "Admin → Users → New → enter Purchase Tester email & password",                      "User created"),
    ("User Mgmt",        "Create Tester 4 account",       "Admin → Users → New → enter Inventory Tester email & password",                     "User created"),
    ("Roles",            "Assign Sales roles → Tester 2", "Users → Tester 2 → assign: Customers, CustomerGroups, CustomerCategories, CustomerContacts, SalesOrders, SalesReports, SalesReturns, DeliveryOrders", "Roles saved"),
    ("Roles",            "Assign Purchase roles → Tester 3","Users → Tester 3 → assign: Vendors, VendorGroups, VendorCategories, VendorContacts, PurchaseOrders, PurchaseReports, PurchaseReturns, GoodsReceives","Roles saved"),
    ("Roles",            "Assign Inventory roles → Tester 4","Users → Tester 4 → assign: Products, ProductGroups, Warehouses, UnitMeasures, TransferOuts, TransferIns, PositiveAdjustments, NegativeAdjustments, Scrappings, StockCounts, TransactionReports, StockReports, MovementReports","Roles saved"),
    ("Settings",         "Configure Company info",        "Settings → My Company → fill name, address, contact → Save",                        "Company info saved"),
    ("Settings",         "Create Tax rate",               "Settings → Tax → New → Name: VAT, Rate: 10% → Save",                                "Tax appears in order dropdowns"),
    ("Settings",         "Verify Number Sequences",       "Settings → Number Sequences → check SO and PO prefixes exist",                      "Sequences present with correct format"),
    ("Login Test",       "Tester 2 login",                "Log in with Tester 2 credentials",                                                  "Only Sales menu items visible"),
    ("Login Test",       "Tester 3 login",                "Log in with Tester 3 credentials",                                                  "Only Purchase menu items visible"),
    ("Login Test",       "Tester 4 login",                "Log in with Tester 4 credentials",                                                  "Only Inventory menu items visible"),
]

P2 = [
    ("Unit Measures",    "Create 'Pieces'",               "Inventory → Unit Measures → New → Name: Pieces",                                    "Unit created"),
    ("Unit Measures",    "Create 'Kg'",                   "Create unit Name: Kg",                                                              "Unit created"),
    ("Unit Measures",    "Create 'Box'",                  "Create unit Name: Box",                                                             "Unit created"),
    ("Unit Measures",    "Edit unit measure",             "Edit 'Box' → change description → Save",                                            "Change saved"),
    ("Unit Measures",    "Delete unit measure",           "Delete 'Box' → confirm",                                                            "Unit no longer in list (soft deleted)"),
    ("Product Groups",   "Create product group",          "Inventory → Product Groups → New → Name: TEST-Group-01",                            "Group created"),
    ("Product Groups",   "Edit product group",            "Edit TEST-Group-01 → rename → Save",                                                "Change saved"),
    ("Products",         "Create TEST-Product-01",        "Inventory → Products → New → fill name, group, unit, price",                        "Product created"),
    ("Products",         "Create TEST-Product-02",        "Create second product",                                                             "Product created"),
    ("Products",         "Create TEST-Product-03",        "Create third product",                                                              "Product created"),
    ("Products",         "Upload product image",          "Edit TEST-Product-01 → upload image → Save",                                        "Image visible on product record"),
    ("Products",         "Edit product price",            "Edit TEST-Product-01 → change price → Save",                                        "Price updated"),
    ("Products",         "Delete product",                "Create a temp product → delete it",                                                 "Not in list, not in order dropdowns"),
    ("Warehouses",       "Create Main Warehouse",         "Inventory → Warehouses → New → Name: TEST-Main-Warehouse",                          "Warehouse created"),
    ("Warehouses",       "Create Secondary Warehouse",    "New → Name: TEST-Secondary-Warehouse",                                             "Warehouse created"),
    ("Warehouses",       "Edit warehouse",                "Edit TEST-Main-Warehouse → update description → Save",                              "Change saved"),
    ("Warehouses",       "Verify 6 system warehouses",    "Check list for: Customer, Vendor, Transfer, Adjustment, StockCount, Scrapping",      "All 6 present"),
    ("Warehouses",       "Try delete system warehouse",   "Attempt to delete 'Customer' system warehouse",                                     "Blocked — error message shown"),
]

P3 = [
    ("Customer Mgmt",    "Create Customer Group",         "Sales → Customer Groups → New → TEST-CustGroup-01",                                 "Group created"),
    ("Customer Mgmt",    "Create Customer Category",      "Sales → Customer Categories → New → TEST-CustCat-01 → link to group",               "Category created"),
    ("Customer Mgmt",    "Create Customer",               "Sales → Customers → New → TEST-Customer-01 → link group & category",                "Customer created"),
    ("Customer Mgmt",    "Create Customer Contact",       "Sales → Customer Contacts → New → link to TEST-Customer-01",                        "Contact created"),
    ("Customer Mgmt",    "Edit Customer",                 "Edit TEST-Customer-01 → change address → Save",                                     "Change saved"),
    ("Sales Orders",     "Create Sales Order",            "Sales → Sales Orders → New → select TEST-Customer-01 → add 2 products with qty → apply VAT tax", "SO created with auto-number SO-XXXX"),
    ("Sales Orders",     "Verify SO status = Draft",      "Check Status field on new SO",                                                      "Status = Draft"),
    ("Sales Orders",     "Confirm Sales Order",           "Click Confirm on the SO",                                                           "Status changes from Draft"),
    ("Sales Orders",     "Generate SO PDF",               "Click PDF button → download",                                                       "PDF downloads with correct data"),
    ("Delivery Orders",  "Create Delivery Order",         "Inventory → Delivery Orders → New → link to confirmed SO → set qty",                "DO created"),
    ("Delivery Orders",  "Confirm Delivery Order",        "Confirm the DO",                                                                    "DO status = Confirmed"),
    ("Stock Check",      "Verify stock DOWN after DO",    "Inventory → Stock Report → filter TEST-Main-Warehouse → TEST-Product-01",            "Stock decreased by delivered qty"),
    ("Sales Returns",    "Create Sales Return",           "Inventory → Sales Returns → New → link to DO",                                      "Return created"),
    ("Sales Returns",    "Confirm Sales Return",          "Confirm the return",                                                                "Return confirmed"),
    ("Stock Check",      "Verify stock UP after return",  "Check Stock Report again",                                                          "Stock back to pre-delivery level"),
    ("Sales Reports",    "Verify SO in Sales Report",     "Sales → Sales Reports → filter by date range covering test SO",                      "SO appears with correct amounts & tax"),
    ("Edge Case",        "Deliver more than ordered",     "Create DO with qty > SO qty",                                                       "System blocks or shows warning"),
    ("Edge Case",        "Delete confirmed SO",           "Try to delete a confirmed Sales Order",                                             "Error shown, delete prevented"),
]

P4 = [
    ("Vendor Mgmt",      "Create Vendor Group",           "Purchase → Vendor Groups → New → TEST-VendGroup-01",                                "Group created"),
    ("Vendor Mgmt",      "Create Vendor Category",        "Purchase → Vendor Categories → New → TEST-VendCat-01",                              "Category created"),
    ("Vendor Mgmt",      "Create Vendor",                 "Purchase → Vendors → New → TEST-Vendor-01 → link group & category",                 "Vendor created"),
    ("Vendor Mgmt",      "Create Vendor Contact",         "Purchase → Vendor Contacts → New → link to TEST-Vendor-01",                         "Contact created"),
    ("Vendor Mgmt",      "Edit Vendor",                   "Edit TEST-Vendor-01 → change details → Save",                                       "Change saved"),
    ("Purchase Orders",  "Create Purchase Order",         "Purchase → Purchase Orders → New → select TEST-Vendor-01 → add 2 products → apply VAT", "PO created with auto-number PO-XXXX"),
    ("Purchase Orders",  "Verify PO status = Draft",      "Check Status field on new PO",                                                      "Status = Draft"),
    ("Purchase Orders",  "Confirm Purchase Order",        "Click Confirm",                                                                     "Status changes"),
    ("Purchase Orders",  "Generate PO PDF",               "Click PDF → download",                                                              "PDF downloads correctly"),
    ("Goods Receives",   "Create Goods Receive",          "Inventory → Goods Receives → New → link to confirmed PO → set qty",                 "GR created"),
    ("Goods Receives",   "Confirm Goods Receive",         "Confirm the GR",                                                                    "GR confirmed"),
    ("Stock Check",      "Verify stock UP after GR",      "Inventory → Stock Report → TEST-Main-Warehouse → TEST-Product-01",                  "Stock increased by received qty"),
    ("Purchase Returns", "Create Purchase Return",        "Purchase → Purchase Returns → New → link to GR",                                   "Return created"),
    ("Purchase Returns", "Confirm Purchase Return",       "Confirm the return",                                                                "Return confirmed"),
    ("Stock Check",      "Verify stock DOWN after return","Check Stock Report",                                                                "Stock decreased back"),
    ("Purchase Reports", "Verify PO in Purchase Report",  "Purchase → Purchase Reports → filter by date",                                      "PO appears with correct amounts"),
    ("Edge Case",        "Receive more than ordered",     "Create GR with qty > PO qty",                                                       "System blocks or warns"),
    ("Edge Case",        "Delete confirmed PO",           "Try to delete confirmed PO",                                                        "Error shown, delete prevented"),
]

P5 = [
    ("Transfer Out",     "Create Transfer Out",           "Inventory → Transfer Out → New → From: TEST-Main → Product: TEST-Product-01 → Qty: 5", "Transfer Out created"),
    ("Transfer Out",     "Confirm Transfer Out",          "Confirm the Transfer Out",                                                          "Status = Confirmed"),
    ("Stock Check",      "Stock DOWN in source warehouse","Stock Report → TEST-Main-Warehouse → TEST-Product-01",                              "Stock decreased by 5"),
    ("Transfer In",      "Create Transfer In",            "Inventory → Transfer In → New → To: TEST-Secondary → link to Transfer Out",         "Transfer In created"),
    ("Transfer In",      "Confirm Transfer In",           "Confirm the Transfer In",                                                           "Status = Confirmed"),
    ("Stock Check",      "Stock UP in dest warehouse",    "Stock Report → TEST-Secondary-Warehouse → TEST-Product-01",                         "Stock increased by 5"),
    ("Transaction Rpt",  "Verify both transfers appear",  "Transaction Report → filter by product",                                            "Transfer Out and Transfer In both listed"),
    ("Pos Adjustment",   "Create Positive Adjustment",    "Inventory → Positive Adjustments → New → TEST-Product-01 → Qty: 10 → TEST-Main",    "Adjustment created"),
    ("Pos Adjustment",   "Confirm Positive Adjustment",   "Confirm",                                                                           "Confirmed"),
    ("Stock Check",      "Stock UP by 10",                "Stock Report",                                                                      "Stock +10"),
    ("Neg Adjustment",   "Create Negative Adjustment",    "Inventory → Negative Adjustments → New → TEST-Product-01 → Qty: 5",                 "Adjustment created"),
    ("Neg Adjustment",   "Confirm Negative Adjustment",   "Confirm",                                                                           "Confirmed"),
    ("Stock Check",      "Stock DOWN by 5",               "Stock Report",                                                                      "Stock -5"),
    ("Edge Case",        "Negative adj > available stock","Create Negative Adjustment for qty > current stock",                                "System warns or blocks"),
    ("Scrapping",        "Create Scrapping",              "Inventory → Scrapping → New → TEST-Product-02 → Qty: 2",                            "Scrapping created"),
    ("Scrapping",        "Confirm Scrapping",             "Confirm",                                                                           "Confirmed"),
    ("Stock Check",      "Stock DOWN after scrapping",    "Stock Report → TEST-Product-02",                                                    "Stock -2"),
    ("Stock Count",      "Create Stock Count",            "Inventory → Stock Count → New → Warehouse: TEST-Main",                              "Stock Count created"),
    ("Stock Count",      "Enter counted quantities",      "Enter qty: match system for P-01, differ for P-02",                                 "Quantities entered"),
    ("Stock Count",      "Confirm Stock Count",           "Confirm",                                                                           "Adjustments recorded for differences"),
    ("Stock Check",      "Stock Report matches count",    "Stock Report → TEST-Main-Warehouse",                                                "Quantities equal the counted values"),
]

P6 = [
    ("Sales Report",     "All SOs appear in report",      "Sales → Sales Reports → date range covering test period",                           "All test SOs listed"),
    ("Sales Report",     "SO amounts correct",            "Cross-check total amounts vs what was entered in SO",                               "Amounts match"),
    ("Purchase Report",  "All POs appear in report",      "Purchase → Purchase Reports → date range covering test period",                     "All test POs listed"),
    ("Purchase Report",  "PO amounts correct",            "Cross-check totals vs PO data",                                                     "Amounts match"),
    ("Stock Report",     "Final stock per warehouse",     "Inventory → Stock Report → filter TEST warehouses",                                 "Stock levels correct after all transactions"),
    ("Stock Report",     "Manual cross-check",            "Sum all In movements - Out movements per product manually → compare to report",      "Report matches manual calculation"),
    ("Transaction Rpt",  "All movements listed",          "Transaction Report → full date range",                                              "Every delivery, receipt, transfer, adjustment, scrapping present"),
    ("Transaction Rpt",  "In/Out types correct",          "Check each row has correct type (In or Out)",                                       "Types are correct"),
    ("Movement Report",  "Product movement summary",      "Inventory → Movement Reports → view per product",                                   "Correct total In, total Out, net per product"),
]

P7 = [
    ("Role Filter",      "Create single-role user",       "Admin → Users → New → assign ONLY 'SalesOrders' role",                             "User created"),
    ("Role Filter",      "Login as single-role user",     "Login with single-role user account",                                               "Menu shows ONLY Sales Orders item"),
    ("Role Filter",      "Other menus hidden",            "Verify Products, Warehouses etc. not in menu",                                     "Hidden items not visible"),
    ("Unauth Access",    "Direct URL blocked",            "Manually type /Products/ProductList in browser while logged as single-role user",   "Access blocked — error or redirect to login"),
    ("No Role",          "Create zero-role user",         "Create user → assign NO roles → Save",                                             "User created"),
    ("No Role",          "Login as zero-role user",       "Login with no-role account",                                                        "Empty menu — no pages accessible"),
    ("Role Update",      "Add role to user",              "Admin → Users → add 'Products' role to single-role user",                          "Role saved"),
    ("Role Update",      "Verify new menu after re-login","Log out and in again as that user",                                                 "Products now visible in menu"),
    ("Role Update",      "Remove role",                   "Admin → Users → remove 'Products' role",                                           "Role removed"),
    ("Role Update",      "Menu updated after re-login",   "Log out and in again",                                                             "Products no longer in menu"),
]

P8 = [
    ("My Profile",       "Update name",                   "Profiles → My Profile → change First/Last name → Save",                            "Name updated in header"),
    ("My Profile",       "Upload avatar",                 "My Profile → upload picture → Save",                                               "Avatar shows in top-right header"),
    ("My Profile",       "Change password",               "My Profile → Change Password → enter current + new password → Save",               "Password changed"),
    ("My Profile",       "Login with new password",       "Logout → login with new password",                                                 "Login successful"),
    ("Company",          "Update company info",           "Settings → My Company → update name, address, logo → Save",                        "Changes saved"),
    ("Company",          "Company shown on PDF",          "Generate any SO or PO PDF",                                                         "Company name/address in PDF header"),
    ("Num Sequences",    "Edit SO prefix",                "Settings → Number Sequences → edit Sales Order prefix to 'TST-SO'",                 "Prefix updated"),
    ("Num Sequences",    "New SO uses updated prefix",    "Create a new Sales Order",                                                          "SO number uses 'TST-SO' prefix"),
    ("Todos",            "Create Todo",                   "Utilities → Todo → New → Title: Test Todo",                                         "Todo created, visible in list"),
    ("Todo Items",       "Create Todo Item",              "Utilities → Todo Items → New → link to Test Todo",                                  "Item created"),
    ("Todos",            "Delete Todo",                   "Delete Test Todo → confirm",                                                        "Todo removed from list"),
]

STATUS_OPTS   = ["Not Tested", "Pass", "Fail", "Blocked", "N/A"]
SEV_OPTS      = ["Critical", "High", "Medium", "Low"]
BUG_ST_OPTS   = ["Open", "In Progress", "Fixed", "Verified", "Closed", "Won't Fix"]
PRI_OPTS      = ["Must Have", "Should Have", "Nice to Have"]
REQ_ST_OPTS   = ["Pending Review", "Approved", "Rejected", "In Progress", "Completed"]
YESNO_OPTS    = ["Yes", "No", "Pending"]

# ─────────────────────────────────────────────────────────────────────────────
# SHEET BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = NAVY
    ws.freeze_panes = "A1"

    _title(ws, "uStock  UAT  Testing  Tracker", NAVY, row=1, cols=8)
    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y')}   |   Environment: https://ustock.unitymicrofund.com")
    c.font = _font(False, 9, "595959"); c.alignment = _align("center"); ws.row_dimensions[2].height = 16

    # Team
    _section(ws, "  TEAM & RESPONSIBILITIES", 4, cols=8)
    _header_row(ws, 5, ["Person", "Role", "Modules Owned", "Phases"], [14, 16, 52, 28], NAVY)
    team = [
        ("Tester 1", "Test Lead",         "Auth · User Mgmt · Settings · Dashboard",                                               "P1, P7, P8 + Final Sign-off"),
        ("Tester 2", "Sales Tester",      "Customers · Sales Orders · Delivery · Sales Returns · Sales Reports",                   "P1, P3, P6"),
        ("Tester 3", "Purchase Tester",   "Vendors · Purchase Orders · Goods Receives · Purchase Returns · Purchase Reports",      "P1, P4, P6"),
        ("Tester 4", "Inventory Tester",  "Products · Warehouses · Transfers · Adjustments · Scrapping · Stock Count · Reports",   "P1, P2, P5, P6"),
    ]
    for i, row_data in enumerate(team):
        r = 6 + i
        bg = LGRAY if i % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            _data_cell(ws, r, col, val, bg)
        ws.row_dimensions[r].height = 28

    # Timeline
    _section(ws, "  TIMELINE  (3 days)", 11, cols=8)
    _header_row(ws, 12, ["Day", "Morning", "Afternoon"], [10, 45, 45], NAVY)
    tl = [
        ("Day 1", "Phase 1 — Setup  (all 4 testers together)",                      "Phase 2 — Master Data  +  Phase 3 — Sales Flow  (parallel)"),
        ("Day 2", "Phase 4 — Purchase Flow  +  Phase 5 — Inventory  (parallel)",    "Phase 6 — Reports  +  Phase 7 — Access Control"),
        ("Day 3", "Bug fixes by dev  +  Retesting failed cases",                    "Phase 8 — Settings  +  Final Sign-off"),
    ]
    for i, row_data in enumerate(tl):
        r = 13 + i
        bg = LGRAY if i % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            _data_cell(ws, r, col, val, bg)
        ws.row_dimensions[r].height = 22

    # Phase summary
    _section(ws, "  PHASE SUMMARY", 17, cols=8)
    _header_row(ws, 18, ["Phase", "Description", "Owner", "Tests", "Status"], [8, 50, 18, 8, 14], NAVY)
    all_phases = [
        ("P1", "Setup & Configuration",                   "All Testers", len(P1)),
        ("P2", "Master Data (Products, Warehouses)",      "Tester 4",    len(P2)),
        ("P3", "Sales Flow  SO → Delivery → Return",      "Tester 2",    len(P3)),
        ("P4", "Purchase Flow  PO → GR → Return",         "Tester 3",    len(P4)),
        ("P5", "Inventory Movements",                     "Tester 4",    len(P5)),
        ("P6", "Reports Verification",                    "All Testers", len(P6)),
        ("P7", "Access Control & Role Testing",           "Tester 1",    len(P7)),
        ("P8", "Profile, Settings & Todos",               "Tester 1",    len(P8)),
    ]
    total_tests = 0
    for i, (code, desc, owner, cnt) in enumerate(all_phases):
        r = 19 + i; total_tests += cnt
        bg = LGRAY if i % 2 == 0 else WHITE
        for col, val in enumerate([code, desc, owner, cnt, "Pending"], 1):
            c = _data_cell(ws, r, col, val, bg, align="center" if col in [1,4,5] else "left")
        ws.row_dimensions[r].height = 22
    # Total
    r = 19 + len(all_phases)
    ws.merge_cells(f"A{r}:C{r}")
    _data_cell(ws, r, 1, "TOTAL", NAVY, bold=True, align="center")
    ws.cell(row=r, column=1).font = _font(True, 11, WHITE)
    _data_cell(ws, r, 4, total_tests, NAVY, bold=True, align="center")
    ws.cell(row=r, column=4).font = _font(True, 11, WHITE)
    ws.row_dimensions[r].height = 22

    # Sign-off criteria note
    _section(ws, "  SIGN-OFF CRITERIA", r+2, cols=8)
    criteria = [
        "✓  Zero Critical bugs open",
        "✓  Zero High bugs open  (or accepted with documented workaround)",
        "✓  Full cycle tested:  PO → Goods Receive → SO → Delivery → Stock Report shows correct numbers",
        "✓  All 4 testers sign off on their phase sheets",
        "✓  Test Lead (Tester 1) gives final approval on Sign-Off sheet",
    ]
    for j, txt in enumerate(criteria):
        rr = r + 3 + j
        ws.merge_cells(f"A{rr}:H{rr}")
        c = ws.cell(row=rr, column=1, value=txt)
        c.fill = _fill(LGRAY if j % 2 == 0 else WHITE)
        c.font = _font(False, 10); c.alignment = _align("left"); c.border = _border()
        ws.row_dimensions[rr].height = 20


def build_checklist(wb, tab_name, code, title, tester, cases):
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = PHASE_CLR[code]
    ws.freeze_panes = "A4"
    bg = PHASE_CLR[code]

    _title(ws, f"uStock UAT  —  {title}    [ Owner: {tester} ]", bg, row=1, cols=8)

    # Status legend
    legend = [("Not Tested","D9D9D9"),("Pass","C6EFCE"),("Fail","FFC7CE"),("Blocked","FFEB9C"),("N/A","F2F2F2")]
    ws.cell(row=2, column=1, value="STATUS LEGEND →").font = _font(True, 9, "595959")
    ws.cell(row=2, column=1).fill = _fill(WHITE)
    ws.cell(row=2, column=1).alignment = _align("right")
    for i, (lbl, lbg) in enumerate(legend, 2):
        c = ws.cell(row=2, column=i, value=lbl)
        c.fill = _fill(lbg); c.font = _font(True, 9, "000000")
        c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[2].height = 18

    headers = ["#", "Category", "Test Case", "Steps to Reproduce", "Expected Result", "Actual Result", "Status", "Notes"]
    widths  = [4,   16,          28,          40,                   34,                34,               13,        22]
    _header_row(ws, 3, headers, widths, bg)

    for i, (cat, tc, steps, expected) in enumerate(cases, 1):
        r = i + 3
        row_bg = LGRAY if i % 2 == 0 else WHITE
        for col, val in enumerate([i, cat, tc, steps, expected, "", "Not Tested", ""], 1):
            _data_cell(ws, r, col, val, row_bg)
        ws.row_dimensions[r].height = 50

    last = len(cases) + 3
    _dropdown(ws, "G", 4, last + 30, STATUS_OPTS)
    ws.sheet_view.showGridLines = True


def build_bug_tracker(wb):
    ws = wb.create_sheet("Bug Tracker")
    ws.sheet_properties.tabColor = PHASE_CLR["BUG"]
    ws.freeze_panes = "A4"

    _title(ws, "uStock UAT  —  Bug Tracker", PHASE_CLR["BUG"], row=1, cols=14)

    # Severity legend
    sev_legend = [("Critical","C00000",WHITE),("High","FF5050",WHITE),("Medium","FFC000","000000"),("Low","70AD47",WHITE)]
    ws.cell(row=2, column=1, value="SEVERITY →").font = _font(True, 9, "595959")
    ws.cell(row=2, column=1).fill = _fill(WHITE)
    ws.cell(row=2, column=1).alignment = _align("right")
    for i, (lbl, lbg, lfg) in enumerate(sev_legend, 2):
        c = ws.cell(row=2, column=i, value=lbl)
        c.fill = _fill(lbg); c.font = _font(True, 9, lfg)
        c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[2].height = 18

    headers = ["#", "Date", "Tester", "Module", "Page / URL",
               "Steps to Reproduce", "Expected", "Actual",
               "Severity", "Assigned To", "Status",
               "Fixed In", "Verified By", "Notes"]
    widths  = [4,   12,     14,       14,       26,
               44,                   34,          34,
               12,         16,            12,
               12,         14,             24]
    _header_row(ws, 3, headers, widths, PHASE_CLR["BUG"])

    for r in range(4, 104):
        row_bg = LGRAY if r % 2 == 0 else WHITE
        for col in range(1, 15):
            _data_cell(ws, r, col, "", row_bg)
        ws.row_dimensions[r].height = 45

    _dropdown(ws, "I",  4, 103, SEV_OPTS)
    _dropdown(ws, "K",  4, 103, BUG_ST_OPTS)


def build_requirements_log(wb):
    ws = wb.create_sheet("Requirements Log")
    ws.sheet_properties.tabColor = PHASE_CLR["REQ"]
    ws.freeze_panes = "A3"

    _title(ws, "uStock UAT  —  New Requirements Log", PHASE_CLR["REQ"], row=1, cols=11)

    headers = ["#", "Date", "Requested By", "Module",
               "Description", "Priority",
               "Effort Estimate", "Decision",
               "Target Release", "Status", "Notes"]
    widths  = [4,   12,     16,             14,
               50,           12,
               16,               22,
               16,               14,       26]
    _header_row(ws, 2, headers, widths, PHASE_CLR["REQ"])

    for r in range(3, 53):
        row_bg = LGRAY if r % 2 == 0 else WHITE
        for col in range(1, 12):
            _data_cell(ws, r, col, "", row_bg)
        ws.row_dimensions[r].height = 45

    _dropdown(ws, "F",  3, 52, PRI_OPTS)
    _dropdown(ws, "J",  3, 52, REQ_ST_OPTS)


def build_signoff(wb):
    ws = wb.create_sheet("Sign-Off")
    ws.sheet_properties.tabColor = PHASE_CLR["SIGN"]
    ws.freeze_panes = "A3"

    _title(ws, "uStock UAT  —  Final Sign-Off Sheet", PHASE_CLR["SIGN"], row=1, cols=10)

    headers = ["Phase", "Description", "Owner",
               "Total", "Passed", "Failed", "Blocked",
               "Sign-Off Date", "Approved", "Comments"]
    widths  = [8,       38,        18,
               8,        8,        8,        9,
               16,                 12,         30]
    _header_row(ws, 2, headers, widths, PHASE_CLR["SIGN"])

    phases = [
        ("P1",    "Setup & Configuration",              "All Testers", len(P1)),
        ("P2",    "Master Data",                        "Tester 4",    len(P2)),
        ("P3",    "Sales Flow",                         "Tester 2",    len(P3)),
        ("P4",    "Purchase Flow",                      "Tester 3",    len(P4)),
        ("P5",    "Inventory Movements",                "Tester 4",    len(P5)),
        ("P6",    "Reports Verification",               "All Testers", len(P6)),
        ("P7",    "Access Control",                     "Tester 1",    len(P7)),
        ("P8",    "Profile & Settings",                 "Tester 1",    len(P8)),
        ("FINAL", "Overall UAT Sign-Off",               "Tester 1 (Test Lead)", "ALL"),
    ]

    for i, (code, desc, owner, total) in enumerate(phases):
        r = 3 + i
        bg = LGRAY if i % 2 == 0 else WHITE
        is_final = code == "FINAL"
        final_bg = "1F3864" if is_final else bg
        final_fg = WHITE    if is_final else "000000"
        for col, val in enumerate([code, desc, owner, total, "", "", "", "", "Pending", ""], 1):
            c = _data_cell(ws, r, col, val, final_bg if is_final else bg,
                           bold=is_final,
                           align="center" if col in [1,4,5,6,7,9] else "left")
            if is_final:
                c.font = _font(True, 10, WHITE)
        ws.row_dimensions[r].height = 30

    _dropdown(ws, "I", 3, 3 + len(phases), YESNO_OPTS)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    print("📋  Building sheets …")
    build_overview(wb)

    sheet_defs = [
        ("P1 - Setup",          "P1", "Phase 1: Setup & Configuration",       "All Testers", P1),
        ("P2 - Master Data",    "P2", "Phase 2: Master Data",                 "Tester 4",    P2),
        ("P3 - Sales Flow",     "P3", "Phase 3: Sales Flow",                  "Tester 2",    P3),
        ("P4 - Purchase Flow",  "P4", "Phase 4: Purchase Flow",               "Tester 3",    P4),
        ("P5 - Inventory",      "P5", "Phase 5: Inventory Movements",         "Tester 4",    P5),
        ("P6 - Reports",        "P6", "Phase 6: Reports Verification",        "All Testers", P6),
        ("P7 - Access Control", "P7", "Phase 7: Access Control",              "Tester 1",    P7),
        ("P8 - Settings",       "P8", "Phase 8: Profile & Settings",          "Tester 1",    P8),
    ]

    for tab, code, title, tester, cases in sheet_defs:
        print(f"   ✓  {tab}")
        build_checklist(wb, tab, code, title, tester, cases)

    print("   ✓  Bug Tracker")
    build_bug_tracker(wb)

    print("   ✓  Requirements Log")
    build_requirements_log(wb)

    print("   ✓  Sign-Off")
    build_signoff(wb)

    out = "uStock_UAT_Testing.xlsx"
    wb.save(out)

    total = sum(len(c) for _, _, _, _, c in sheet_defs)
    print(f"\n✅  Done!  →  {out}")
    print(f"   Sheets  : 12  (Overview + 8 Phase checklists + Bug Tracker + Requirements Log + Sign-Off)")
    print(f"   Test cases: {total}")
    print(f"\n📤  To use in Google Sheets:")
    print(f"   1. Go to sheets.google.com")
    print(f"   2. File → Import → Upload → select {out}")
    print(f"   3. Choose 'Replace spreadsheet' → Import")

if __name__ == "__main__":
    main()
